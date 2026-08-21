from __future__ import annotations

import base64
from datetime import UTC, date, datetime, time
from decimal import Decimal
from io import BytesIO
from typing import Any
from uuid import UUID

import sqlalchemy as sa
from sqlalchemy.orm import Session, undefer

from app.core.fechas import TZ_LOCAL
from app.db.models import (
    AjusteCaja,
    Cheque,
    Cliente,
    Compensacion,
    CompensacionImputacion,
    Cuota,
    DeudaSimple,
    Fiado,
    GastoOperativo,
    MovimientoCaja,
    MovimientoEfectivo,
    Pasivo,
    Prestamo,
)

BACKUP_VERSION = 1
ALEMBIC_REVISION = "0013"

# ── Columnas por tabla ──────────────────────────────────────────────────────

# Marca de anulación (migración 0017). Va en toda tabla anulable: si el backup no
# la llevara, un ciclo export→import resucitaría los registros anulados y la caja
# volvería a descuadrar.
_ANUL = ["anulado_at", "anulado_por", "motivo_anulacion"]

_CL = ["id", "nombre", "cuit", "telefono", "created_at", "updated_at"]
_CH = [
    "id", "nro_cheque", "banco", "monto", "fecha_emision", "fecha_pago",
    "porcentaje_compra", "porcentaje_venta", "ganancia", "estado",
    # Cuánto se abonó al comprarlo (§Comprar sin abonar). Sin este dato el import
    # devuelve un cheque comprado a deber como pagado, y al editarlo el resync le
    # asienta el egreso entero que nunca salió de la caja.
    "monto_abonado",
    "ultimo_evento_manual_at", "ultimo_operador_id", "ultimo_motivo_manual",
    "foto", "foto_mime", "cliente_origen_id", "cliente_destino_id",
    # Marca de cartera preexistente (§Apertura). Sin ella el import devuelve esos
    # cheques como compras normales y al editarlos se les asienta el egreso
    # COMPRA_CHEQUE que el régimen de apertura quita: la plata se descuenta dos veces.
    "es_carga_inicial",
    "created_at", "updated_at", *_ANUL,
]
_PR = [
    "id", "cliente_id", "credito", "moneda", "cuotas", "frecuencia",
    "total_a_cobrar", "ganancia", "estado", "fecha_inicio", "created_at", "updated_at",
    *_ANUL,
]
_CU = [
    "id", "prestamo_id", "numero_cuota", "fecha_vencimiento", "monto", "monto_pagado",
    "estado", "fecha_cobro", "created_at", "updated_at",
]
_MO = [
    "id", "cliente_id", "tipo", "moneda", "monto", "cotizacion_aplicada",
    "ganancia", "usd_restante", "fecha_operacion", "observaciones", "created_at", "updated_at",
    # Cuánto se abonó de la compra (§Comprar sin abonar): igual que en los cheques,
    # sin esto una compra a deber vuelve como pagada y se gana un egreso al editarla.
    "monto_abonado",
    # Las dos marcas tienen que viajar: un lote de apertura o de ajuste que vuelve
    # como compra normal se gana líneas de caja que nunca existieron al editarlo,
    # y `_rehacer_lote_usd` deja de encontrarlo para reemplazarlo.
    "es_apertura", "es_ajuste",
    *_ANUL,
]
_FI = [
    "id", "cheque_id", "cliente_id", "monto_original", "porcentaje_venta",
    "saldo_pendiente", "estado", "fecha_fiado", "created_at", "updated_at", *_ANUL,
]
_PA = [
    "id", "acreedor", "concepto", "monto", "saldo_pendiente", "moneda",
    # Cotización default de los pagos que cruzan monedas (§5).
    "cotizacion_pago",
    # De qué compra salió el pasivo (§Comprar sin abonar). Sin el vínculo, anular
    # esa compra deja de encontrar su deuda y queda viva plata que ya no se debe.
    "origen_tipo", "origen_id",
    # Si con la deuda entró plata al cajón y qué día (§5). Sin esto, un ciclo
    # export→import convierte un préstamo recibido en deuda común y le borra el
    # ingreso al reconstruir la caja.
    "ingreso_caja", "fecha_ingreso",
    "estado", "fecha_vencimiento", "fecha_cancelacion", "observaciones",
    "created_at", "updated_at", *_ANUL,
]
_GA = [
    "id", "concepto", "monto", "moneda", "fecha_operacion", "hora_operacion",
    "observaciones", "created_at", "updated_at", *_ANUL,
]
_DS = [
    "id", "cliente_id", "concepto", "monto", "saldo_pendiente", "moneda",
    "estado", "fecha", "fecha_cancelacion", "observaciones", "cotizacion_pago",
    "created_at", "updated_at", *_ANUL,
]
_MC = [
    "id", "fecha", "moneda", "tipo", "categoria", "monto", "ganancia",
    # Con qué medio se pagó cada pasivo y el $/USD de los pagos que cruzan monedas
    # (migración 0014). Sin ellos el import los devuelve en NULL y se pierde el
    # detalle de auditoría de esas líneas.
    "medio_pago", "cotizacion",
    "referencia_tipo", "referencia_id", "detalle", "created_at", "updated_at",
]
_AJ = [
    "id", "fecha", "moneda", "tipo", "motivo", "monto", "cotizacion_usd",
    "lote_id", "descripcion", "operador_id", "created_at", "updated_at", *_ANUL,
]
_CO = [
    "id", "fecha", "cliente_id", "acreedor", "moneda", "monto", "moneda_deuda",
    "moneda_pasivo", "cotizacion", "imputado_cliente", "imputado_pasivo",
    # El excedente y el pasivo que se le creó al cliente por él: sin el vínculo,
    # revertir la compensación no encuentra qué anular y esa plata le queda a
    # favor para siempre.
    "excedente", "pasivo_excedente_id",
    "observaciones", "created_at", "updated_at", *_ANUL,
]
_CI = [
    "id", "compensacion_id", "entidad_tipo", "entidad_id", "monto", "cancelo",
    "created_at",
]

# ── Validación de schema ────────────────────────────────────────────────────

_REQUIRED: dict[str, frozenset[str]] = {
    "clientes":             frozenset({"id", "nombre"}),
    "cheques":              frozenset({"id", "nro_cheque", "monto", "estado", "fecha_emision"}),
    "prestamos":            frozenset({"id", "cliente_id", "credito", "moneda", "cuotas", "frecuencia", "total_a_cobrar", "estado"}),
    "cuotas":               frozenset({"id", "prestamo_id", "numero_cuota", "fecha_vencimiento", "monto", "estado"}),
    "movimientos_efectivo": frozenset({"id", "tipo", "moneda", "monto", "cotizacion_aplicada"}),
    "fiados":               frozenset({"id", "cheque_id", "cliente_id", "monto_original", "saldo_pendiente", "estado"}),
    "pasivos":              frozenset({"id", "acreedor", "concepto", "monto", "moneda", "estado"}),
    "gastos_operativos":    frozenset({"id", "concepto", "monto", "moneda", "fecha_operacion"}),
    # movimientos_caja y deudas_simples son opcionales: los backups anteriores no
    # los traen y deben seguir importando. Solo se validan si la tabla viene presente.
    "movimientos_caja":     frozenset({"id", "fecha", "moneda", "tipo", "categoria", "monto"}),
    "deudas_simples":       frozenset({"id", "cliente_id", "concepto", "monto", "saldo_pendiente", "moneda", "estado", "fecha"}),
    "ajustes_caja":         frozenset({"id", "fecha", "moneda", "tipo", "motivo", "monto"}),
    "compensaciones":       frozenset({"id", "fecha", "cliente_id", "acreedor", "moneda", "monto", "moneda_deuda"}),
    "compensacion_imputaciones": frozenset({"id", "compensacion_id", "entidad_tipo", "entidad_id", "monto"}),
}


def _validate_schema(tablas: dict) -> list[str]:
    errors: list[str] = []
    for tbl, required in _REQUIRED.items():
        for i, row in enumerate(tablas.get(tbl, [])):
            missing = required - set(row.keys())
            if missing:
                errors.append(
                    f"`{tbl}` fila {i}: faltan campos: {', '.join(sorted(missing))}"
                )
    return errors


# ── Tipos para deserialización ──────────────────────────────────────────────

_UUID_COLS = frozenset({
    "id", "cliente_id", "cheque_id", "prestamo_id",
    "cliente_origen_id", "cliente_destino_id", "referencia_id", "lote_id",
})
_DEC_COLS = frozenset({
    "monto", "monto_pagado", "credito", "total_a_cobrar", "ganancia",
    "porcentaje_compra", "porcentaje_venta", "cotizacion_aplicada",
    "monto_original", "saldo_pendiente", "usd_restante", "cotizacion_pago",
    "cotizacion_usd", "cotizacion",
    # Compras a deber y compensaciones (§Comprar sin abonar, §Compensación).
    "monto_abonado", "imputado_cliente", "imputado_pasivo", "excedente",
})
_DT_COLS = frozenset({"created_at", "updated_at", "ultimo_evento_manual_at", "anulado_at"})
_BYTES_COLS = frozenset({"foto"})


# ── Serialización ───────────────────────────────────────────────────────────

def _out(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, UUID):
        return str(v)
    if isinstance(v, Decimal):
        return str(v)
    if isinstance(v, bytes):
        return base64.b64encode(v).decode()
    if hasattr(v, "isoformat"):
        return v.isoformat()
    if hasattr(v, "value"):
        return v.value
    return v


def _serialize(row: Any, cols: list[str]) -> dict:
    return {c: _out(getattr(row, c)) for c in cols}


def exportar_json(db: Session) -> dict:
    cheques = db.query(Cheque).options(undefer(Cheque.foto)).all()
    return {
        "version": BACKUP_VERSION,
        "exported_at": datetime.now(tz=UTC).isoformat(),
        "alembic_revision": ALEMBIC_REVISION,
        "tablas": {
            "clientes":             [_serialize(r, _CL) for r in db.query(Cliente).all()],
            "cheques":              [_serialize(r, _CH) for r in cheques],
            "prestamos":            [_serialize(r, _PR) for r in db.query(Prestamo).all()],
            "cuotas":               [_serialize(r, _CU) for r in db.query(Cuota).all()],
            "movimientos_efectivo": [_serialize(r, _MO) for r in db.query(MovimientoEfectivo).all()],
            "fiados":               [_serialize(r, _FI) for r in db.query(Fiado).all()],
            "pasivos":              [_serialize(r, _PA) for r in db.query(Pasivo).all()],
            "gastos_operativos":    [_serialize(r, _GA) for r in db.query(GastoOperativo).all()],
            "deudas_simples":       [_serialize(r, _DS) for r in db.query(DeudaSimple).all()],
            "movimientos_caja":     [_serialize(r, _MC) for r in db.query(MovimientoCaja).all()],
            "ajustes_caja":         [_serialize(r, _AJ) for r in db.query(AjusteCaja).all()],
            "compensaciones":       [_serialize(r, _CO) for r in db.query(Compensacion).all()],
            # El detalle por renglón viaja aparte: es lo que permite revertir una
            # compensación devolviendo exactamente lo que sacó de cada deuda.
            "compensacion_imputaciones": [
                _serialize(r, _CI) for r in db.query(CompensacionImputacion).all()
            ],
        },
    }


# ── Deserialización ─────────────────────────────────────────────────────────

def _cv(
    row: dict,
    date_cols: frozenset[str] = frozenset(),
    dt_extra: frozenset[str] = frozenset(),
    time_cols: frozenset[str] = frozenset(),
) -> dict:
    r: dict = {}
    for k, v in row.items():
        if v is None:
            r[k] = None
        elif k in _UUID_COLS:
            r[k] = UUID(v) if isinstance(v, str) else v
        elif k in _DEC_COLS:
            r[k] = Decimal(str(v))
        elif k in _DT_COLS or k in dt_extra:
            r[k] = datetime.fromisoformat(v) if isinstance(v, str) else v
        elif k in date_cols:
            r[k] = date.fromisoformat(v) if isinstance(v, str) else v
        elif k in time_cols:
            r[k] = time.fromisoformat(v) if isinstance(v, str) else v
        elif k in _BYTES_COLS:
            r[k] = base64.b64decode(v) if isinstance(v, str) else v
        else:
            r[k] = v
    return r


_DATE_CH = frozenset({"fecha_emision", "fecha_pago"})
_DATE_PR = frozenset({"fecha_inicio"})
_DATE_CU = frozenset({"fecha_vencimiento", "fecha_cobro"})
_DATE_FI = frozenset({"fecha_fiado"})
_DATE_PA = frozenset({"fecha_vencimiento", "fecha_cancelacion"})
_DATE_GA = frozenset({"fecha_operacion"})
_TIME_GA = frozenset({"hora_operacion"})
_DT_MO   = frozenset({"fecha_operacion"})
_DATE_MC = frozenset({"fecha"})
_DATE_DS = frozenset({"fecha", "fecha_cancelacion"})
_DATE_AJ = frozenset({"fecha"})
_DATE_CO = frozenset({"fecha"})


def importar_json(db: Session, data: dict) -> dict[str, int]:
    if data.get("version") != BACKUP_VERSION:
        raise ValueError(
            f"Versión de backup incompatible: {data.get('version')!r}. "
            f"Se esperaba {BACKUP_VERSION}."
        )

    tablas = data.get("tablas", {})
    if not isinstance(tablas, dict):
        raise ValueError("El backup no contiene una clave 'tablas' válida.")

    errors = _validate_schema(tablas)
    if errors:
        preview = "\n".join(errors[:10])
        suffix = f"\n… y {len(errors) - 10} más." if len(errors) > 10 else ""
        raise ValueError(f"{len(errors)} error(es) de schema:\n{preview}{suffix}")

    try:
        # `ajustes_caja` va primero: referencia a `movimientos_efectivo` por su lote.
        for tbl in (
            "compensacion_imputaciones", "compensaciones",
            "ajustes_caja", "movimientos_caja", "cuotas", "fiados", "deudas_simples",
            "movimientos_efectivo", "prestamos", "cheques", "pasivos",
            "gastos_operativos", "clientes",
        ):
            db.execute(sa.text(f"DELETE FROM {tbl}"))  # noqa: S608

        def bulk(model: Any, rows: list[dict]) -> None:
            if rows:
                db.execute(sa.insert(model), rows)

        bulk(Cliente,            [_cv(r) for r in tablas.get("clientes", [])])
        bulk(Cheque,             [_cv(r, date_cols=_DATE_CH) for r in tablas.get("cheques", [])])
        bulk(Prestamo,           [_cv(r, date_cols=_DATE_PR) for r in tablas.get("prestamos", [])])
        bulk(Cuota,              [_cv(r, date_cols=_DATE_CU) for r in tablas.get("cuotas", [])])
        bulk(MovimientoEfectivo, [_cv(r, dt_extra=_DT_MO) for r in tablas.get("movimientos_efectivo", [])])
        bulk(Fiado,              [_cv(r, date_cols=_DATE_FI) for r in tablas.get("fiados", [])])
        bulk(Pasivo,             [_cv(r, date_cols=_DATE_PA) for r in tablas.get("pasivos", [])])
        bulk(GastoOperativo,     [_cv(r, date_cols=_DATE_GA, time_cols=_TIME_GA) for r in tablas.get("gastos_operativos", [])])
        bulk(DeudaSimple,        [_cv(r, date_cols=_DATE_DS) for r in tablas.get("deudas_simples", [])])
        bulk(MovimientoCaja,     [_cv(r, date_cols=_DATE_MC) for r in tablas.get("movimientos_caja", [])])
        # Después de movimientos_efectivo: cada ajuste en USD apunta a su lote.
        bulk(AjusteCaja,         [_cv(r, date_cols=_DATE_AJ) for r in tablas.get("ajustes_caja", [])])
        # Después de clientes y pasivos: la compensación apunta a los dos.
        bulk(Compensacion,       [_cv(r, date_cols=_DATE_CO) for r in tablas.get("compensaciones", [])])
        bulk(CompensacionImputacion, [_cv(r) for r in tablas.get("compensacion_imputaciones", [])])

        db.commit()
    except Exception:
        db.rollback()
        raise

    tabla_names = (
        "clientes", "cheques", "prestamos", "cuotas",
        "movimientos_efectivo", "fiados", "pasivos", "gastos_operativos",
        "deudas_simples", "movimientos_caja", "ajustes_caja",
        "compensaciones", "compensacion_imputaciones",
    )
    return {t: len(tablas.get(t, [])) for t in tabla_names}


# ── Excel ───────────────────────────────────────────────────────────────────

def _fmt_excel(v: Any) -> Any:
    if v is None:
        return ""
    if isinstance(v, UUID):
        return str(v)
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, bytes):
        return ""
    if isinstance(v, datetime):
        return v.replace(tzinfo=None)
    if isinstance(v, time):
        return v.strftime("%H:%M:%S")
    if hasattr(v, "value"):
        return v.value
    return v


def _xl_image(foto: bytes) -> Any:
    """Redimensiona y convierte bytes de foto a imagen embebible en openpyxl."""
    try:
        from PIL import Image as PILImage
        from openpyxl.drawing.image import Image as XLImage

        img = PILImage.open(BytesIO(foto))
        img.thumbnail((110, 110), PILImage.Resampling.LANCZOS)
        buf = BytesIO()
        img.save(buf, format="PNG")
        buf.seek(0)
        xl = XLImage(buf)
        xl.width, xl.height = 110, 110
        return xl
    except Exception:
        return None


def exportar_excel(
    db: Session,
    *,
    desde: date | None = None,
    hasta: date | None = None,
    tablas_incluidas: list[str] | None = None,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    def _inc(name: str) -> bool:
        return tablas_incluidas is None or name in tablas_incluidas

    def _date_filter(q: Any, model: Any) -> Any:
        # Los timestamps se guardan en UTC, pero el operador elige fechas en hora
        # local (Argentina). Convertimos el día local completo a UTC para no
        # traspapelar operaciones nocturnas al día equivocado (igual que reportes.py).
        if desde:
            desde_dt = datetime.combine(desde, time.min, tzinfo=TZ_LOCAL).astimezone(UTC)
            q = q.filter(model.created_at >= desde_dt)
        if hasta:
            hasta_dt = datetime.combine(hasta, time.max, tzinfo=TZ_LOCAL).astimezone(UTC)
            q = q.filter(model.created_at <= hasta_dt)
        return q

    def _vivos(q: Any, model: Any) -> Any:
        """Excluye las operaciones anuladas.

        Solo aplica al Excel: es un reporte de trabajo, y una operación dada de
        baja no debe figurar como si se hubiera hecho. El backup JSON, en cambio,
        SÍ las conserva con su marca — es una copia fiel de la base."""
        return q.filter(model.anulado_at.is_(None))

    HEADER_FILL = PatternFill("solid", fgColor="4F46E5")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
    ALT_FILL    = PatternFill("solid", fgColor="F1F5F9")

    wb = Workbook()
    wb.remove(wb.active)  # type: ignore[arg-type]

    def write_headers(ws: Any, headers: list[str]) -> None:
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 20

    def add_sheet(title: str, headers: list[str], data_rows: list[list]) -> None:
        ws = wb.create_sheet(title)
        write_headers(ws, headers)
        for ri, row in enumerate(data_rows, 2):
            fill = ALT_FILL if ri % 2 == 0 else None
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                if fill:
                    cell.fill = fill
        for ci in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 20

    # ── Clientes ────────────────────────────────────────────────────────────
    if _inc("clientes"):
        clientes = _date_filter(db.query(Cliente), Cliente).all()
        add_sheet(
            "Clientes",
            ["ID", "Nombre", "CUIT", "Teléfono", "Creado"],
            [[_fmt_excel(r.id), r.nombre, r.cuit, r.telefono, _fmt_excel(r.created_at)]
             for r in clientes],
        )

    # ── Cheques (foto embebida) ──────────────────────────────────────────────
    if _inc("cheques"):
        cheques = _vivos(
            _date_filter(db.query(Cheque).options(undefer(Cheque.foto)), Cheque), Cheque
        ).all()
        ws_ch = wb.create_sheet("Cheques")
        ch_headers = [
            "ID", "Nro Cheque", "Banco", "Monto", "Fecha Emisión", "Fecha Pago",
            "% Compra", "% Venta", "Ganancia", "Estado",
            "Cliente Origen ID", "Cliente Destino ID", "Creado", "Foto",
        ]
        write_headers(ws_ch, ch_headers)
        foto_col = len(ch_headers)
        for ri, r in enumerate(cheques, 2):
            data_vals = [
                _fmt_excel(r.id), r.nro_cheque, r.banco, _fmt_excel(r.monto),
                _fmt_excel(r.fecha_emision), _fmt_excel(r.fecha_pago),
                _fmt_excel(r.porcentaje_compra), _fmt_excel(r.porcentaje_venta),
                _fmt_excel(r.ganancia), _fmt_excel(r.estado),
                _fmt_excel(r.cliente_origen_id), _fmt_excel(r.cliente_destino_id),
                _fmt_excel(r.created_at),
            ]
            fill = ALT_FILL if ri % 2 == 0 else None
            for ci, val in enumerate(data_vals, 1):
                cell = ws_ch.cell(row=ri, column=ci, value=val)
                if fill:
                    cell.fill = fill
            if r.foto:
                xl_img = _xl_image(r.foto)
                if xl_img:
                    xl_img.anchor = f"{get_column_letter(foto_col)}{ri}"
                    ws_ch.add_image(xl_img)
                    ws_ch.row_dimensions[ri].height = 85
                else:
                    ws_ch.cell(row=ri, column=foto_col, value="[error al procesar imagen]")
            else:
                ws_ch.cell(row=ri, column=foto_col, value="—")
        for ci in range(1, len(ch_headers) + 1):
            ws_ch.column_dimensions[get_column_letter(ci)].width = 17 if ci == foto_col else 20

    # ── Préstamos ────────────────────────────────────────────────────────────
    if _inc("prestamos"):
        prestamos = _vivos(_date_filter(db.query(Prestamo), Prestamo), Prestamo).all()
        add_sheet(
            "Préstamos",
            ["ID", "Cliente ID", "Crédito", "Moneda", "Cuotas", "Frecuencia",
             "Total a Cobrar", "Ganancia", "Estado", "Fecha Inicio", "Creado"],
            [[
                _fmt_excel(r.id), _fmt_excel(r.cliente_id),
                _fmt_excel(r.credito), _fmt_excel(r.moneda),
                r.cuotas, _fmt_excel(r.frecuencia),
                _fmt_excel(r.total_a_cobrar), _fmt_excel(r.ganancia),
                _fmt_excel(r.estado), _fmt_excel(r.fecha_inicio), _fmt_excel(r.created_at),
            ] for r in prestamos],
        )

    # ── Cuotas ───────────────────────────────────────────────────────────────
    if _inc("cuotas"):
        # Las cuotas no tienen marca propia: heredan la del préstamo al que
        # pertenecen, así que se excluyen las de préstamos anulados.
        cuotas = (
            _date_filter(db.query(Cuota), Cuota)
            .join(Cuota.prestamo)
            .filter(Prestamo.anulado_at.is_(None))
            .all()
        )
        add_sheet(
            "Cuotas",
            ["ID", "Préstamo ID", "Nro Cuota", "Vencimiento", "Monto", "Pagado", "Estado",
             "Fecha Cobro", "Creado"],
            [[
                _fmt_excel(r.id), _fmt_excel(r.prestamo_id), r.numero_cuota,
                _fmt_excel(r.fecha_vencimiento), _fmt_excel(r.monto), _fmt_excel(r.monto_pagado),
                _fmt_excel(r.estado), _fmt_excel(r.fecha_cobro), _fmt_excel(r.created_at),
            ] for r in cuotas],
        )

    # ── Movimientos ──────────────────────────────────────────────────────────
    if _inc("movimientos_efectivo"):
        movimientos = _vivos(
            _date_filter(db.query(MovimientoEfectivo), MovimientoEfectivo), MovimientoEfectivo
        ).all()
        add_sheet(
            "Movimientos",
            ["ID", "Cliente ID", "Tipo", "Moneda", "Monto", "Cotización",
             "Ganancia", "Stock USD restante", "Fecha Operación", "Observaciones", "Creado"],
            [[
                _fmt_excel(r.id), _fmt_excel(r.cliente_id), _fmt_excel(r.tipo),
                _fmt_excel(r.moneda), _fmt_excel(r.monto), _fmt_excel(r.cotizacion_aplicada),
                _fmt_excel(r.ganancia), _fmt_excel(r.usd_restante), _fmt_excel(r.fecha_operacion),
                r.observaciones, _fmt_excel(r.created_at),
            ] for r in movimientos],
        )

    # ── Fiados ───────────────────────────────────────────────────────────────
    if _inc("fiados"):
        fiados = _vivos(_date_filter(db.query(Fiado), Fiado), Fiado).all()
        add_sheet(
            "Fiados",
            ["ID", "Cheque ID", "Cliente ID", "Monto Original", "% Venta",
             "Saldo Pendiente", "Estado", "Fecha Fiado", "Creado"],
            [[
                _fmt_excel(r.id), _fmt_excel(r.cheque_id), _fmt_excel(r.cliente_id),
                _fmt_excel(r.monto_original), _fmt_excel(r.porcentaje_venta),
                _fmt_excel(r.saldo_pendiente), _fmt_excel(r.estado),
                _fmt_excel(r.fecha_fiado), _fmt_excel(r.created_at),
            ] for r in fiados],
        )

    # ── Pasivos ──────────────────────────────────────────────────────────────
    if _inc("pasivos"):
        pasivos = _vivos(_date_filter(db.query(Pasivo), Pasivo), Pasivo).all()
        add_sheet(
            "Pasivos",
            ["ID", "Acreedor", "Concepto", "Monto", "Saldo Pendiente", "Moneda",
             "Estado", "Vencimiento", "Cancelación", "Observaciones", "Creado"],
            [[
                _fmt_excel(r.id), r.acreedor, r.concepto,
                _fmt_excel(r.monto), _fmt_excel(r.saldo_pendiente), _fmt_excel(r.moneda),
                _fmt_excel(r.estado), _fmt_excel(r.fecha_vencimiento),
                _fmt_excel(r.fecha_cancelacion), r.observaciones, _fmt_excel(r.created_at),
            ] for r in pasivos],
        )

    # ── Gastos Operativos ─────────────────────────────────────────────────────
    if _inc("gastos_operativos"):
        gastos = _vivos(_date_filter(db.query(GastoOperativo), GastoOperativo), GastoOperativo).all()
        add_sheet(
            "Gastos Operativos",
            ["ID", "Concepto", "Monto", "Moneda", "Fecha Operación",
             "Hora Operación", "Observaciones", "Creado"],
            [[
                _fmt_excel(r.id), r.concepto, _fmt_excel(r.monto), _fmt_excel(r.moneda),
                _fmt_excel(r.fecha_operacion), _fmt_excel(r.hora_operacion),
                r.observaciones, _fmt_excel(r.created_at),
            ] for r in gastos],
        )

    # ── Caja (libro de movimientos) ───────────────────────────────────────────
    if _inc("movimientos_caja"):
        mov_caja = _date_filter(db.query(MovimientoCaja), MovimientoCaja).all()
        add_sheet(
            "Caja",
            ["ID", "Fecha", "Moneda", "Tipo", "Categoría", "Monto", "Ganancia",
             "Referencia", "Detalle", "Creado"],
            [[
                _fmt_excel(r.id), _fmt_excel(r.fecha), _fmt_excel(r.moneda),
                _fmt_excel(r.tipo), _fmt_excel(r.categoria), _fmt_excel(r.monto),
                _fmt_excel(r.ganancia), r.referencia_tipo, r.detalle, _fmt_excel(r.created_at),
            ] for r in mov_caja],
        )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
